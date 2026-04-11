"""
Excel（日常点検表2024 .xlsm）から点検項目をインポートするスクリプト

【Excelファイルの構造】
  シート「点検表」を解析します。
  シートは左右2ブロック構成です：
    左ブロック (A〜G列, index 0〜6):
      col 0 (A): フロア（例: RF, 2F, B1）※建物名ヘッダー行時は建物名
      col 1 (B): 点検場所名
      col 2 (C): 点検番号（例: 1-1）
      col 3 (D): '・'（データ行の目印）
      col 4 (E): 判定結果
      col 5 (F): メモ
    右ブロック (H〜N列, index 7〜13):
      col  7 (H): フロア ※建物名ヘッダー行時は建物名
      col  8 (I): 点検場所名
      col  9 (J): 点検番号
      col 10 (K): '・'（データ行の目印）
      col 11 (L): 判定結果
      col 12 (M): メモ

  建物名ヘッダー行の判定:
    - A列 or H列に建物名（文字列）
    - C列 or J列が空（コードなし）
    - D列 or K列が空（'・'なし）

  データ行の判定:
    - D列 or K列 == '・'
    - C列 or J列にコード（例: "1-1"）
    - B列 or I列に点検場所名

【使い方】
  python import_excel.py [Excelファイルパス]

  パス省略時は同フォルダの "日常点検表2024 .xlsm" を自動検索します。

  例:
    python import_excel.py
    python import_excel.py "日常点検表2024 .xlsm"
    python import_excel.py "..\01_tenken\日常点検表2024.xlsm"
"""

import json
import sqlite3
import sys
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("エラー: openpyxl が必要です。pip install openpyxl を実行してください。")
    sys.exit(1)

DB_PATH         = Path(__file__).parent / "tenken.db"
EXCEL_DIR       = Path(__file__).parent
SHEET_NAME      = "点検表"
FAULT_SHEET     = "故障リスト・機器運転替え"

# ─── 建物名 → 曜日マッピング ──────────────────────────────────
# 週番号情報（第1週）は建物名から自動抽出するため、ここでは除去済みの名前を使う
BUILDING_DAY_MAP = {
    "沈砂池棟":           1,   # 月曜日
    "処理水再利用設備":    2,   # 火曜日
    "塩素接触棟":         2,   # 火曜日
    "汚泥調整槽":         2,   # 火曜日
    "管廊（雑排水ポンプ）": 2,  # 火曜日
    "第二ポンプ施設":      3,   # 水曜日
    "第三ポンプ施設":      3,   # 水曜日
    "特高棟":            3,   # 水曜日
    "滞水池":            3,   # 水曜日
    "中央管理棟":         3,   # 水曜日
    "ホッパー棟":         3,   # 水曜日
    "二次処理棟":         4,   # 木曜日
    "第5系列水処理施設":   4,   # 木曜日
    "1～4系水処理施設":   4,   # 木曜日
    "7・8系水処理棟":     5,   # 金曜日
    "7・8系送風機棟":     5,   # 金曜日
    "7・8系管廊・流量計室": 5,  # 金曜日
}


def extract_week_filter(building_name: str):
    """建物名から週フィルタを抽出（例: '第二ポンプ施設（第1週）' → 1）"""
    m = re.search(r'第([1-4])週', building_name)
    return int(m.group(1)) if m else None


def normalize_building(raw_name: str) -> str:
    """
    建物名を正規化する
      - 週情報「（第N週）」を除去
      - 汚泥調整槽のサブセクション（"汚泥調整槽10・20" 等）を親名に統一
    """
    name = re.sub(r'（第[1-4]週）', '', raw_name).strip()
    if name.startswith("汚泥調整槽"):
        name = "汚泥調整槽"
    return name


def get_day_of_week(building_raw: str) -> int:
    """建物名（生）から曜日番号を返す。不明なら -1"""
    normalized = normalize_building(building_raw)
    if normalized in BUILDING_DAY_MAP:
        return BUILDING_DAY_MAP[normalized]
    # 前方一致フォールバック
    for key, day in BUILDING_DAY_MAP.items():
        if normalized.startswith(key) or key in normalized:
            return day
    return -1


def _add_footnote(footnotes: dict, building: str, text: str):
    """※N のテキストを footnotes[building][※N] に格納する"""
    if not building:
        return
    text = text.strip()
    m = re.match(r'^(※\d+)[\s\u3000]*(.*)', text, re.DOTALL)
    if m:
        ref     = m.group(1)          # "※1"
        content = m.group(2).strip()  # 本文
        footnotes.setdefault(building, {})[ref] = content


def _resolve_refs(memo: str, fn_map: dict) -> str:
    """memo 中の ※N を fn_map の実テキストに置換する"""
    def replace(m):
        return fn_map.get(m.group(0), m.group(0))
    return re.sub(r'※\d+', replace, memo)


def detect_dot_columns(rows: list) -> list[int]:
    """
    行データから '・' が2件以上出現する列インデックス一覧を返す（昇順）。
    Excelの列追加・移動に対応するための自動検出。
    検出できない場合はデフォルト [3, 10]（D列・K列）を返す。
    """
    from collections import Counter
    counter = Counter()
    for row in rows:
        for i, v in enumerate(row):
            if v == "・":
                counter[i] += 1
    cols = sorted(col for col, cnt in counter.items() if cnt >= 2)
    return cols if cols else [3, 10]


# ─── 先発データ解析ヘルパー ───────────────────────────────────

def parse_senpatu_formula(s: str):
    """
    =IF(MOD($P$12,N)=...) 形式の数式を12か月分のリストに変換する。
    解析できない場合は None を返す。
    """
    # 2択: =IF(MOD($P$12,2)=0,"even","odd")
    m = re.match(r'^=IF\(MOD\(\$P\$12,2\)=0,"([^"]+)","([^"]+)"\)$', s)
    if m:
        v_even, v_odd = m.group(1), m.group(2)
        return [v_even if month % 2 == 0 else v_odd for month in range(1, 13)]

    # 3択: =IF(MOD($P$12,3)=0,"v0",IF(MOD($P$12,3)=1,"v1","v2"))
    m = re.match(
        r'^=IF\(MOD\(\$P\$12,3\)=0,"([^"]+)",IF\(MOD\(\$P\$12,3\)=1,"([^"]+)","([^"]+)"\)\)$',
        s
    )
    if m:
        v0, v1, v2 = m.group(1), m.group(2), m.group(3)
        def pick(month):
            r = month % 3
            return v0 if r == 0 else (v1 if r == 1 else v2)
        return [pick(month) for month in range(1, 13)]

    return None


def _resolve_and_parse(ws, formula_str, depth=0):
    """セル参照を辿りながら数式を解析して12か月リストを返す。失敗時はNone。"""
    if depth > 5 or not formula_str:
        return None
    s = str(formula_str).strip()
    if not s.startswith("="):
        return None

    result = parse_senpatu_formula(s)
    if result is not None:
        return result

    # セル参照（例: =G73）の解決
    m = re.match(r'^=([A-Z]+\d+)$', s)
    if m:
        ref = m.group(1)
        try:
            ref_val = ws[ref].value
            if ref_val:
                return _resolve_and_parse(ws, ref_val, depth + 1)
        except Exception:
            pass

    return None


def collect_senpatu_data(ws_formula) -> tuple:
    """
    数式ワークブックのシートから先発データを収集する。

    Returns:
        item_senpatu : dict  {点検番号コード -> 先発値文字列}
        groups       : list  senpatu_groups テーブル用レコードのリスト
    """
    item_senpatu: dict = {}

    # ── '・' 列位置を自動検出してブロック定義を構築 ──────────────
    # 先発列は '・' の3列右（元: G=D+3, N=K+3）
    _dot_cols = detect_dot_columns([[c.value for c in r] for r in ws_formula.iter_rows()])
    SENPATU_BLOCKS = [{"dot": dc, "code": dc - 1, "sen": dc + 3} for dc in _dot_cols]

    for row_cells in ws_formula.iter_rows(values_only=False):
        for cols in SENPATU_BLOCKS:
            def sval(idx):
                return row_cells[idx].value if idx < len(row_cells) else None

            dot_val  = sval(cols["dot"])
            code_val = sval(cols["code"])

            if dot_val != "・" or not code_val:
                continue

            code_str = str(code_val).strip()
            sen_raw  = sval(cols["sen"])

            if sen_raw is None:
                item_senpatu[code_str] = ""
                continue

            sen_str = str(sen_raw).strip()

            if sen_str in ("A", "B", "C"):
                item_senpatu[code_str] = sen_str
            elif sen_str.startswith("="):
                monthly = _resolve_and_parse(ws_formula, sen_str)
                item_senpatu[code_str] = json.dumps(monthly, ensure_ascii=False) if monthly else ""
            else:
                item_senpatu[code_str] = sen_str

    # ── 先発グループテーブルを収集 ──────────────────────────────
    # (day_of_week, group_name, row_start, row_end, name_col_0idx, num_col_0idx)
    GROUP_DEFS = [
        (1, "A", 34,  40,  8, 10),   # 月: I=name(8), K=num(10)
        (2, "A", 84,  87,  8, 10),   # 火A: H84-H87
        (2, "B", 88,  90,  8, 10),   # 火B: H88-H90
        (3, "A", 124, 134, 1,  3),   # 水A: B=name(1), D=num(3)
        (3, "B", 135, 136, 1,  3),   # 水B
        (4, "A", 225, 227, 8, 10),   # 木A
        (4, "B", 228, 233, 8, 10),   # 木B
        (4, "C", 234, 236, 8, 10),   # 木C
        (5, "A", 274, 278, 8, 10),   # 金A
        (5, "B", 279, 280, 8, 10),   # 金B
    ]

    groups = []
    for day, grp, r_start, r_end, nc, kc in GROUP_DEFS:
        sort_order = 0
        for row_num in range(r_start, r_end + 1):
            name_val = ws_formula.cell(row=row_num, column=nc + 1).value
            num_val  = ws_formula.cell(row=row_num, column=kc + 1).value

            if name_val is None:
                continue
            name_str = str(name_val).strip()
            if not name_str:
                continue

            monthly = None
            if num_val:
                ns = str(num_val).strip()
                if ns.startswith("="):
                    monthly = _resolve_and_parse(ws_formula, ns)
                if monthly is None:
                    monthly = [ns] * 12

            if monthly is None:
                monthly = [""] * 12

            sort_order += 1
            groups.append({
                "day_of_week":     day,
                "group_name":      grp,
                "sort_order":      sort_order,
                "machine_name":    name_str,
                "monthly_numbers": json.dumps(monthly, ensure_ascii=False),
            })

    return item_senpatu, groups


def collect_fault_data(wb) -> dict:
    """
    「故障リスト・機器運転替え」シートから {点検番号: 故障内容} を返す。
    3行目のヘッダー行で「No.」列（点検番号）と「故障内容」列を検出し、
    4行目以降のデータを読み込む。列位置が変わっても自動対応する。
    シートが存在しない場合は空辞書を返す。
    """
    if FAULT_SHEET not in wb.sheetnames:
        print(f"  注意: シート '{FAULT_SHEET}' が見つかりません。故障データはスキップします。")
        return {}

    ws = wb[FAULT_SHEET]
    fault_map: dict = {}

    # ── 3行目ヘッダーから「No.」列と「故障内容」列を検出 ──────────
    max_col = ws.max_column or 20
    header = [ws.cell(row=3, column=c).value for c in range(1, max_col + 1)]

    # 「No.」「№」どちらの表記も対応
    _NO_LABELS = {"No.", "No", "NO.", "NO", "№"}
    code_cols    = [i + 1 for i, v in enumerate(header)
                    if v is not None and str(v).strip() in _NO_LABELS]
    content_cols = [i + 1 for i, v in enumerate(header)
                    if v is not None and "故障内容" in str(v).strip()]

    # 各「No.」列の右側で最初の「故障内容」列とペアリング
    block_pairs = []
    for cc in code_cols:
        for kc in content_cols:
            if kc > cc:
                block_pairs.append((cc, kc))
                break

    if not block_pairs:
        # フォールバック: 従来の固定位置（B/E, J/M）
        print("  注意: 故障シートのヘッダーが検出できません。デフォルト列（B→E, J→M）を使用します。")
        block_pairs = [(2, 5), (10, 13)]
    else:
        pairs_str = ", ".join(f"No.=列{cc}→故障内容=列{kc}" for cc, kc in block_pairs)
        print(f"  故障シート列検出: {pairs_str}")

    # ── 4行目以降のデータを読み込む ────────────────────────────
    max_row = ws.max_row or 50
    for col_code, col_content in block_pairs:
        for row in range(4, max_row + 1):
            code_val    = ws.cell(row=row, column=col_code).value
            content_val = ws.cell(row=row, column=col_content).value
            if code_val and content_val:
                code_str    = str(code_val).strip()
                content_str = str(content_val).strip()
                if code_str and content_str:
                    fault_map[code_str] = content_str

    return fault_map


def is_building_header(name_val, code_val, dot_val) -> bool:
    """建物名ヘッダー行かどうかを判定"""
    if not isinstance(name_val, str) or not name_val.strip():
        return False
    skip = {'場所', '備考', '先発切替機器', '共通', 'A', 'B',
            'センター長', '管理担当係長', '水再生担当係長', '主任'}
    if name_val.strip() in skip:
        return False
    if code_val is not None:
        return False
    if dot_val is not None:
        return False
    return True


def parse_sheet(ws) -> list:
    """
    シートを解析して点検項目のリストを返す。
    '・' 列位置を自動検出することで、列の追加・移動に対応する。
    """
    # 全行を先読みして '・' 列位置を自動検出
    all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
    dot_cols = detect_dot_columns(all_rows)
    block_names = ["left", "right"] + [f"block{i}" for i in range(2, len(dot_cols))]
    BLOCKS = {
        name: {"floor": dc - 3, "loc": dc - 2, "code": dc - 1,
               "dot": dc, "result": dc + 1, "memo": dc + 2}
        for name, dc in zip(block_names, dot_cols)
    }
    _detected = ", ".join(f"{n}=列{c['dot']+1}" for n, c in BLOCKS.items())
    print(f"  列自動検出: {_detected}")

    def gv(row, idx):
        """安全に列の値を取得（負インデックス・範囲外は None）"""
        return row[idx] if 0 <= idx < len(row) else None


    items    = []
    code_detected_count = 0  # '・' なしでコードパターン検出により取り込んだ件数
    footnotes: dict = {}  # {building_name: {"※1": "text", ...}}

    # 各ブロックの状態管理
    state = {
        side: {
            "building":    None,   # 現在の建物名（生）
            "week_filter": None,   # 週フィルタ
            "floor":       None,   # 現在のフロア（引き継ぎ）
            "last_loc":    None,   # マージセル対応: 直前の有効な場所名
            "sort_order":  0,      # 建物内の表示順
            "in_footnote": False,  # 備考（脚注）セクション中か
        }
        for side in BLOCKS
    }

    for row in all_rows:

        for side, cols in BLOCKS.items():
            st = state[side]
            floor_val = gv(row, cols["floor"])
            loc_val   = gv(row, cols["loc"])   # 左=B, 右=I（脚注テキストも同列）
            code_val  = gv(row, cols["code"])
            dot_val   = gv(row, cols["dot"])
            memo_val  = gv(row, cols["memo"])

            floor_str = str(floor_val).strip() if isinstance(floor_val, str) else ""

            # ── 備考セクション開始（A列またはH列 == '備考'） ────────
            if floor_str == "備考":
                st["in_footnote"] = True
                # 同行にすでに ※N テキストがある場合は収集
                if loc_val:
                    _add_footnote(footnotes, normalize_building(st["building"] or ""), str(loc_val))
                continue

            # ── データ行の判定: '・' があるか、または点検番号パターン（数字-数字）がある ──
            _code_tmp = str(code_val).strip() if code_val is not None else ""
            _is_data_row = (dot_val == "・") or bool(
                re.match(r'^\d+-\d+$', _code_tmp) and st["building"]
            )

            # ── 備考セクション内 ─────────────────────────────────
            if st["in_footnote"]:
                # 新しい建物ヘッダー or データ行が来たら脱出
                if is_building_header(floor_val, code_val, dot_val) or _is_data_row:
                    st["in_footnote"] = False
                    # fall through して通常処理へ
                else:
                    if loc_val:
                        _add_footnote(footnotes, normalize_building(st["building"] or ""), str(loc_val))
                    continue

            # ── 建物名ヘッダー行 ──────────────────────────────────
            if is_building_header(floor_val, code_val, dot_val):
                raw = str(floor_val).strip()
                st["building"]    = raw
                st["week_filter"] = extract_week_filter(raw)
                st["floor"]       = None  # フロア引き継ぎリセット
                st["last_loc"]    = None  # 建物が変わったらリセット
                st["sort_order"]  = 0
                continue

            # ── データ行（'・' または点検番号コード一致） ──────────
            # loc_val が None はマージセルの可能性があるため、直前の有効値を引き継ぐ
            if _is_data_row:
                if dot_val != "・" and code_val:
                    code_detected_count += 1
                if not code_val:
                    print(f"  ⚠ スキップ（コードなし）: building={st['building']!r} loc={loc_val!r} floor={floor_val!r}")
                    continue
                code_str = str(code_val).strip()
                if loc_val is not None:
                    loc_str = str(loc_val).strip()
                    st["last_loc"] = loc_str
                elif st["last_loc"]:
                    loc_str = st["last_loc"]
                else:
                    print(f"  ⚠ スキップ（場所名なし）: code={code_str!r} building={st['building']!r}")
                    continue  # 場所名が特定できない行はスキップ

                # フロア値の更新（値があれば引き継ぐ）
                if isinstance(floor_val, str) and floor_val.strip():
                    st["floor"] = floor_val.strip()

                if st["building"] is None:
                    print(f"  ⚠ スキップ（建物名未確定）: code={code_str!r}")
                    continue  # 建物名が未確定はスキップ

                day = get_day_of_week(st["building"])
                if day < 0:
                    print(f"  ⚠ 未知の建物: {st['building']} (code={code_str}) → スキップ")
                    continue

                # フロア情報を description に保存
                desc = f"{st['floor']}" if st["floor"] else ""

                # F/M列のメモを base_memo として取り込む
                base_memo = str(memo_val).strip() if isinstance(memo_val, str) and memo_val.strip() else ""

                # E/L列の判定内容を result_hint として取り込む
                result_raw = gv(row, cols["result"])
                result_hint = str(result_raw).strip() if result_raw is not None and str(result_raw).strip() else ""

                st["sort_order"] += 1
                items.append({
                    "code":        code_str,
                    "building":    normalize_building(st["building"]),
                    "location":    loc_str,
                    "description": desc,
                    "base_memo":   base_memo,
                    "result_hint": result_hint,
                    "day_of_week": day,
                    "week_filter": st["week_filter"],
                    "sort_order":  st["sort_order"],
                })

    # ── ※N 参照を実テキストに解決 ────────────────────────────────
    for item in items:
        if item["base_memo"] and "※" in item["base_memo"]:
            fn_map = footnotes.get(item["building"], {})
            if fn_map:
                item["base_memo"] = _resolve_refs(item["base_memo"], fn_map)

    if code_detected_count:
        print(f"  注意: '・' なし・コードパターンで検出した行: {code_detected_count} 件")
    return items


def import_from_excel(excel_path: str, no_confirm: bool = False):
    path = Path(excel_path)
    if not path.exists():
        print(f"エラー: ファイルが見つかりません: {path}")
        sys.exit(1)

    print(f"読み込み中: {path.name}")

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, keep_vba=True)
    except Exception as e:
        print(f"Excelの読み込みに失敗しました: {e}")
        sys.exit(1)

    if SHEET_NAME not in wb.sheetnames:
        print(f"エラー: シート '{SHEET_NAME}' が見つかりません。")
        print(f"  利用可能なシート: {wb.sheetnames}")
        wb.close()
        sys.exit(1)

    ws = wb[SHEET_NAME]
    all_items = parse_sheet(ws)
    wb.close()

    if not all_items:
        print("インポートするデータがありません。")
        return

    # ── 先発データを数式ワークブックから収集 ─────────────────────
    item_senpatu: dict = {}
    senpatu_groups_data: list = []
    try:
        wb_f = openpyxl.load_workbook(str(path), data_only=False, keep_vba=True)
        if SHEET_NAME in wb_f.sheetnames:
            ws_f = wb_f[SHEET_NAME]
            item_senpatu, senpatu_groups_data = collect_senpatu_data(ws_f)
            print(f"先発データ: 点検行 {len(item_senpatu)} 件, グループ {len(senpatu_groups_data)} エントリ")
        wb_f.close()
    except Exception as e:
        print(f"  注意: 先発データの読み込みをスキップしました: {e}")

    # 先発値を各点検項目にマージ（未取得の項目は空文字）
    for item in all_items:
        item["senpatu"] = item_senpatu.get(item["code"], "")

    # ── 故障データを data_only ワークブックから収集 ──────────────
    fault_map: dict = {}
    try:
        wb_d = openpyxl.load_workbook(str(path), data_only=True, keep_vba=True)
        fault_map = collect_fault_data(wb_d)
        wb_d.close()
        if fault_map:
            print(f"故障データ: {len(fault_map)} 件")
    except Exception as e:
        print(f"  注意: 故障データの読み込みをスキップしました: {e}")

    # 故障内容を各点検項目にマージ（未取得の項目は空文字）
    for item in all_items:
        item["fault_memo"] = fault_map.get(item["code"], "")

    # ── 曜日別集計を表示 ─────────────────────────────────────
    from collections import defaultdict
    day_names = {1: "月", 2: "火", 3: "水", 4: "木", 5: "金"}
    day_counts = defaultdict(int)
    bldg_counts = defaultdict(int)
    for item in all_items:
        day_counts[item["day_of_week"]] += 1
        bldg_counts[item["building"]] += 1

    print(f"\n取得結果:")
    for day in sorted(day_counts):
        print(f"  {day_names[day]}曜日: {day_counts[day]} 件")
    print(f"  合計: {len(all_items)} 件")
    print(f"\n建物別:")
    for bldg, cnt in sorted(bldg_counts.items(), key=lambda x: x[0]):
        print(f"  {bldg}: {cnt} 件")

    # ── DB更新 ───────────────────────────────────────────────
    if not DB_PATH.exists():
        print(f"\nエラー: DBファイルが見つかりません: {DB_PATH}")
        print("先に 'python database.py' を実行してDBを初期化してください。")
        sys.exit(1)

    if no_confirm:
        print(f"\ninspection_items テーブルを {len(all_items)} 件で上書きします。")
    else:
        confirm = input(f"\ninspection_items テーブルを {len(all_items)} 件で上書きします。よろしいですか？ [y/N]: ")
        if confirm.strip().lower() != "y":
            print("キャンセルしました。")
            return

    conn = sqlite3.connect(str(DB_PATH))
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM inspection_items")
        cur.executemany("""
            INSERT INTO inspection_items
                (code, building, location, description, base_memo, senpatu,
                 fault_memo, result_hint, day_of_week, week_filter, sort_order)
            VALUES
                (:code, :building, :location, :description, :base_memo, :senpatu,
                 :fault_memo, :result_hint, :day_of_week, :week_filter, :sort_order)
        """, all_items)

        # 先発グループを保存
        try:
            cur.execute("DELETE FROM senpatu_groups")
            if senpatu_groups_data:
                cur.executemany("""
                    INSERT INTO senpatu_groups
                        (day_of_week, group_name, sort_order, machine_name, monthly_numbers)
                    VALUES
                        (:day_of_week, :group_name, :sort_order, :machine_name, :monthly_numbers)
                """, senpatu_groups_data)
        except Exception as eg:
            print(f"  注意: senpatu_groups の保存に失敗しました: {eg}")

        conn.commit()
        print(f"\n完了: {len(all_items)} 件の点検項目をインポートしました。")
        if senpatu_groups_data:
            print(f"  先発グループ: {len(senpatu_groups_data)} エントリを保存しました。")
        print("  注意: inspection_results（点検結果）は変更されていません。")
    except Exception as e:
        conn.rollback()
        print(f"DB保存エラー: {e}")
        sys.exit(1)
    finally:
        conn.close()


def find_excel_auto() -> str:
    """同フォルダからExcelファイルを自動検索（更新日付が最新のファイルを返す）"""
    files = []
    for pattern in ["*.xlsm", "*.xlsx"]:
        files.extend(EXCEL_DIR.glob(pattern))
    if not files:
        return None
    latest = max(files, key=lambda f: f.stat().st_mtime)
    return str(latest)


if __name__ == "__main__":
    if len(sys.argv) >= 2:
        target = sys.argv[1]
    else:
        target = find_excel_auto()
        if target:
            print(f"Excelファイルを自動検出: {Path(target).name}")
        else:
            print("エラー: Excelファイルが見つかりません。")
            print("使い方: python import_excel.py <Excelファイルパス>")
            sys.exit(1)

    import_from_excel(target)
