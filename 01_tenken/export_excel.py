"""
点検結果 → Excel書き出しモジュール

Excelシートの「結果」列（左ブロック: D列, 右ブロック: K列）に
点検結果（〇/×/△）を書き込み、BytesIOで返す。
・結果列には '・' が入っており、それを 〇/×/△ で上書きする
・書き込み対象シートは「点検表」
・元のExcelファイルは変更しない（メモリ上でコピーして処理する）
"""

import openpyxl
from io import BytesIO
from pathlib import Path
from datetime import datetime

EXCEL_DIR  = Path(__file__).parent
SHEET_NAME = "点検表"

# ブロック定義（列番号は openpyxl の 1-based）
# 左ブロック (A〜G列)
LEFT_BLOCK = {
    "code_col":   3,   # C: 点検番号
    "dot_col":    4,   # D: '結果'（'・' → 〇/×/△ に上書き）
    "result_col": 4,   # D: 結果書き込み列（dot列と同じ）
    "memo_col":   6,   # F: 備考
}
# 右ブロック (H〜N列)
RIGHT_BLOCK = {
    "code_col":   10,  # J: 点検番号
    "dot_col":    11,  # K: '結果'（'・' → 〇/×/△ に上書き）
    "result_col": 11,  # K: 結果書き込み列（dot列と同じ）
    "memo_col":   13,  # M: 備考
}


def find_excel() -> Path | None:
    """同フォルダからExcelファイルを自動検索"""
    for pattern in ["*.xlsm", "*.xlsx"]:
        files = list(EXCEL_DIR.glob(pattern))
        if files:
            return files[0]
    return None


def build_cell_map(ws) -> dict:
    """
    シートを解析して {点検番号: {row, result_col, memo_col}} のマップを返す。
    左右ブロックを同時処理する。
    """
    cell_map = {}

    def gv(row, col_1based):
        idx = col_1based - 1
        return row[idx] if idx < len(row) else None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row = list(row)

        for blk in (LEFT_BLOCK, RIGHT_BLOCK):
            code_val = gv(row, blk["code_col"])
            dot_val  = gv(row, blk["dot_col"])

            # データ行: dot列が '・' または '-' かつ code列に値あり
            # （月曜は '・'、火〜金曜は '-' をプレースホルダーとして使用）
            if dot_val in ("・", "-") and code_val is not None:
                code_str = str(code_val).strip()
                if code_str:
                    cell_map[code_str] = {
                        "row":        row_idx,
                        "result_col": blk["result_col"],
                        "memo_col":   blk["memo_col"],
                    }
    return cell_map


def export_to_excel(results: list, date_str: str, date_label: str | None = None) -> tuple[bytes, int, str]:
    """
    点検結果をExcelに書き込んでBytesで返す。

    Parameters
    ----------
    results : list of dict
        [{"code": "1-1", "result": "〇", "memo": "..."}, ...]
    date_str : str
        "YYYY-MM-DD" 形式の点検日

    Returns
    -------
    (excel_bytes, written_count, source_filename)
    """
    excel_path = find_excel()
    if excel_path is None:
        raise FileNotFoundError(
            "Excelテンプレートが見つかりません。"
            f"'{EXCEL_DIR}' に .xlsm または .xlsx ファイルを置いてください。"
        )

    # data_only=True でセルマップを構築（数式セルの計算済み値を取得）
    wb_read = openpyxl.load_workbook(str(excel_path), keep_vba=True, data_only=True)
    if SHEET_NAME not in wb_read.sheetnames:
        raise ValueError(f"シート '{SHEET_NAME}' が見つかりません。")
    cell_map = build_cell_map(wb_read[SHEET_NAME])
    wb_read.close()

    # 書き込み用に再読み込み（数式・マクロを保持）
    wb = openpyxl.load_workbook(str(excel_path), keep_vba=True)
    ws = wb[SHEET_NAME]

    # 点検日を日付セルに書き込む（row 5: タイトル行）
    try:
        if date_label is None:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            date_label = f"{dt.month}月{dt.day}日"
        ws.cell(row=5, column=3).value = date_label  # C5
    except Exception:
        pass

    # 結果を書き込む
    result_map = {r["code"]: r for r in results if r.get("result")}
    written = 0

    for code, info in cell_map.items():
        if code not in result_map:
            continue
        r = result_map[code]

        # 結果列（D/K）に書き込む（'・' を 〇/×/△ で上書き）
        ws.cell(row=info["row"], column=info["result_col"]).value = r["result"]

        # メモ列に書き込む（既存メモは上書きしない = 新規分のみ）
        if r.get("memo"):
            existing_memo = ws.cell(row=info["row"], column=info["memo_col"]).value
            if not existing_memo:
                ws.cell(row=info["row"], column=info["memo_col"]).value = r["memo"]

        written += 1

    # メモリ上で保存して返す
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue(), written, excel_path.name
